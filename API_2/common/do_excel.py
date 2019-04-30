# -*- coding: utf-8 -*-
# @Time     :2019/4/17  9:58
# @Author   :liying
# @Email    :1025452202@qq.com
# @File     :do_excel.py

#数据的读和写
import openpyxl
from API_2.common import http_request
class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None
        self.sql=None
class DoExcel:
    def __init__(self,file_name,sheet_name):
        #异常处理
        #try:
            self.file_name=file_name
            self.sheet_name=sheet_name
            self.workbook=openpyxl.load_workbook(file_name)
            self.sheet=self.workbook[sheet_name]
        # except Exception as e:
        #     raise e

    def get_cases(self):
        max_row=self.sheet.max_row#获取最大行数

        cases=[]  #列表，存放所有的测试用例
        for r in range(2,max_row+1):
            # case={}
            # case['case_id']=self.sheet.cell(row=r,column=1)
            # case['title'] = self.sheet.cell(row=r, column=2)
            # cases.append(case)
            case=Case()#实例
            case.case_id=self.sheet.cell(row=r,column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value
            cases.append(case)
            # aaa=[]
            # for a in range(1,7):
            #     asd=self.sheet.cell(row=r,column=a).value
            #     aaa.append(asd)
            # cases.append(aaa)

        self.workbook.close()
        return cases #返回cases列表

    def write_result(self,row,actual,result):
        sheet=self.workbook[self.sheet_name]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()

if __name__ == '__main__':
    sheet_names=['login','recharge','withdraw']
    res=http_request.HTTPRequest2()#同一个session
    for item in sheet_names:
        do_excel=DoExcel('testcase.xlsx',sheet_name=item)
        cases=do_excel.get_cases()


        for case in cases:
            print(case.title)
            print(case.url)
            print(case.data)
            print(type(case.data))
            print(case.__dict__)#所有属性字典

            resp=res.request(case.method,case.url,case.data)
            print(resp.text)
            print(case.expected)
            actual=resp.text

            if case.expected==actual:
                do_excel.write_result(case.case_id+1,actual,'PASS')
            else:
                do_excel.write_result(case.case_id+1,actual,'FAIL')